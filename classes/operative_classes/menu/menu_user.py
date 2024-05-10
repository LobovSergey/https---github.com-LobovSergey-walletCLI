from dataclasses import dataclass
import os
from openpyxl import load_workbook
from openpyxl import Workbook
from pandas import DataFrame
import numpy as np
from numpy.typing import NDArray
from datetime import datetime
from classes.operative_classes.table.main_page_table import MainPage
from conf.var import (
    DB_NAME,
    USERLIST_TITLES,
    USERMENU_CHOISES,
    HISTORY_CHOISES,
    SEARCH_CHOISES,
    SORT_CHOISES,
)


@dataclass
class MenuUser:
    user: str
    balance = int
    """Класс меню пользователя"""

    def _add_value_at_balance(self, value: int, operand: bool) -> None:
        """Изменение баланса на занчение value. operand=True является 'Пополнением',  а False - 'Cписанием'"""
        if operand:
            self.balance += value
        else:
            self.balance -= value

    def _set_balance(self) -> None:
        """Установка баланса, путем суммирования всех операций."""
        data = self._get_frame()
        self.balance = data[USERLIST_TITLES[2]].sum()

    def _get_frame(self) -> DataFrame:
        """Вывод DataFrame для удобного отображения и работы с данными, которые не требуют изменения в таблице."""
        book = self._get_book()
        data = self._get_array(book)
        data_frame = DataFrame(data=data[1:], columns=data[0])
        return data_frame

    def _edit_balance(self, table: MainPage, flag: bool = None) -> None:
        """Логика изменения баланса при Списаниях или Пополнениях"""
        if flag:
            operand = "Пополнение"
        else:
            operand = "Списание"
        while True:
            try:
                value = int(input(f"\n--{operand}--\nВведите сумму:\n"))
                if not flag:
                    if self.balance >= value:
                        break
                else:
                    break
                print("Недопустимая сумма")
            except ValueError:
                print("Некорректный ввод")

        description = input("Введите описание операциии (опционально)\n")
        page = table.table[self.user]
        page.append([datetime.now(), operand, value, description])
        self._add_value_at_balance(value=value, operand=flag)
        table.save()

    def _get_array(self, table: MainPage) -> NDArray:
        """Numpy array для передачи данных в pandas"""
        return np.array([val for val in table[self.user].values])

    def _get_book(self) -> Workbook:
        """Чтение таблицы для работы с pandas"""
        return load_workbook(os.path.abspath(f"{DB_NAME}.xlsx"))

    def _sort_by(self, frame: DataFrame) -> None:
        """Меню сортировки и ее логика. Предлагается отсортировать те столбцы, котоыре присутсвуют в файле.
        (!) Сортировка по возрастанию"""
        while True:
            choise = input(SORT_CHOISES)
            if choise == "0":
                return
            elif choise in ["1", "2", "3"]:
                print(frame.sort_values(USERLIST_TITLES[int(choise) - 1]))
            else:
                print("Некорректный выбор. Попробуйте еще раз")

    def _search_by(self, frame: DataFrame) -> None:
        """Фильтрация данных в датафрейме при полном совпадении запрашиваемого значения и результата. В противном случае выведет пустой датафрейм
        (!) Можно перенастройить под логику .contains() для частичного, а не полного совпадения
        """
        while True:
            choise = input(SEARCH_CHOISES)
            if choise == "0":
                return
            elif choise in ["1", "2", "3"]:
                searched = input("Искомое значение:\n")
                print(
                    frame.loc[frame[USERLIST_TITLES[int(choise) - 1]].isin([searched])]
                )
            else:
                print("Некорректный выбор. Попробуйте еще раз")

    def _edit_operation(self, table: MainPage) -> None:
        """Меню и логика работы с редактированием операции.
        Выбор строки -> Выбор столбца -> Ввод нового значения.
        Тип данных не учитывается.
        После внесения сохраняется в таблице"""
        while True:
            page = table.table[self.user]
            i = -1
            for row in page.values:
                i += 1
                if i == 0:
                    continue
                print(i, end=" ")
                for value in row:
                    print(value, end=" ")
                print()

            choise = input("Введите номер строки: (< - назад)")
            if choise == "<":
                return
            try:
                print("Выберите поле по которому хотите изменить данные")
                for i in range(len(USERLIST_TITLES)):
                    print(f"{i+1} - {USERLIST_TITLES[i]}")
                choise_col = int(input())
                val = input("Введите новое значение:\n")

                cell = page.cell(row=int(choise) + 1, column=choise_col)
                cell.value = val
                table.save()

            except ValueError:
                print("Некорректный выбор. Попробуйте еще раз")
            except KeyError:
                print("Такого столбца нет")

    def _get_histiry(self, table: Workbook) -> None:
        """Меню отображения и работы с вашими операциями"""
        while True:
            data_frame = self._get_frame()
            print(f"--История операций--\n{data_frame}")
            choise = input(HISTORY_CHOISES)
            if choise == "0":
                return
            elif choise == "1":
                self._sort_by(data_frame)
            elif choise == "2":
                self._search_by(data_frame)
            elif choise == "3":
                self._edit_operation(table)
            else:
                print("Некорректный выбор. Попробуйте еще раз")

    def get_usermenu(self, table: Workbook, choise: int = None) -> None:
        """Основное меню вашего личного кабинета."""
        self._set_balance()
        while True:
            print(
                f"\n--Личный кабинет --\nПользователь : {self.user}\nБаланс: {self.balance}"
            )
            choise = input(USERMENU_CHOISES)
            if choise == "0":
                return
            elif choise == "1":
                self._edit_balance(table=table, flag=True)
            elif choise == "2":
                self._edit_balance(table=table, flag=False)
            elif choise == "3":
                self._get_histiry(table)
            else:
                print("Некорректный выбор. Попробуйте еще раз")
